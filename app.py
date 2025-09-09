from flask import Flask, render_template, request, redirect, flash
import os
import re
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
import requests  # <-- NUEVO

app = Flask(__name__)
app.secret_key = 'clave_secreta_para_flash'

# Configuraciones
UPLOAD_FOLDER = r'C:/Users/Administrador/OneDrive - Grupo OLAM Paraguay/Facturas_excel'
EMAIL_EXCEL_FOLDER = os.path.join(UPLOAD_FOLDER, 'correos_excel')
EMAIL_EXCEL_FILE = os.path.join(EMAIL_EXCEL_FOLDER, 'correos.xlsx')
ALLOWED_EXTENSIONS = {'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Funciones auxiliares
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def valid_email(email):
    return re.match(r"[^@]+@[^@]+\.[^@]+", email)

def guardar_correo_en_excel(email):
    os.makedirs(EMAIL_EXCEL_FOLDER, exist_ok=True)

    if not os.path.exists(EMAIL_EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Correos"
        ws.append(["ID", "Correo"])
        wb.save(EMAIL_EXCEL_FILE)
        wb.close()

    wb = load_workbook(EMAIL_EXCEL_FILE)
    ws = wb.active
    new_id = ws.max_row
    ws.append([new_id, email])
    wb.save(EMAIL_EXCEL_FILE)
    wb.close()

# NUEVO: Enviar a servidor local
def enviar_a_servidor_local(email, archivos):
    url_local = 'http://hybrydportalfacturas-app' # REEMPLAZA con la IP o nombre de Hybrid Connection

    files = {}
    for tipo, archivo in archivos.items():
        if archivo and allowed_file(archivo.filename):
            archivo.stream.seek(0)
            files[tipo] = (archivo.filename, archivo.stream, 'application/pdf')

    data = {'email': email}
    
    headers = {'Authorization': 'Bearer 9f82a7f1-2341-456c-b812-9abcde123457'} 

    try:
        response = requests.post(url_local, files=files, data=data, headers=headers)  # <-- Aquí agregás headers
        if response.status_code == 200:
            flash('Datos enviados al servidor local correctamente.')
        else:
            flash(f'Error al enviar datos al servidor local: {response.status_code} - {response.text}')
    except Exception as e:
        flash(f'No se pudo conectar con el servidor local: {str(e)}')


# Ruta principal
@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        email = request.form.get('email')
        factura = request.files.get('factura')
        orden = request.files.get('orden')
        remision = request.files.get('remision')

        if not email or not valid_email(email):
            flash('Correo electrónico no válido.')
            return redirect(request.url)

        if not (factura or orden or remision):
            flash('Debe adjuntar al menos un archivo: Factura, Orden de Compra o Remisión.')
            return redirect(request.url)

        archivos = {
            'Factura': factura,
            'Orden': orden,
            'Remision': remision
        }

        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        archivos_guardados = 0

        for tipo, archivo in archivos.items():
            if archivo and archivo.filename and allowed_file(archivo.filename):
                filename = secure_filename(f"{tipo}_{archivo.filename}")
                saved_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                archivo.save(saved_path)
                archivos_guardados += 1
            elif archivo and archivo.filename:
                flash(f'Archivo no permitido: {archivo.filename}')

        if archivos_guardados == 0:
            flash('Ningún archivo válido fue cargado.')
            return redirect(request.url)

        try:
            guardar_correo_en_excel(email)
        except Exception as e:
            flash(f'Error al guardar el correo en Excel (Azure): {str(e)}')

        try:
            enviar_a_servidor_local(email, archivos)
        except Exception as e:
            flash(f'Error al enviar al servidor local: {str(e)}')

        flash(f'Se recibieron {archivos_guardados} archivo(s) correctamente de {email}.')
        return redirect(request.url)

    return render_template('formulario.html')

if __name__ == '__main__':
    app.run(debug=True)
